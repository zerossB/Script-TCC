from datetime import date

from openpyxl import load_workbook

from .exceptions import XlsxExeption


class Xlsx(object):
    def __init__(self, license):
        self.workbook = "output/%s.xlsx" % license
        self.book = load_workbook(self.workbook)
        self.sheet = self.book.create_sheet(str(date.today()))

    def saveToWorkbook(self, header, data):
        # if len(header[0]) != len(data[0]):
        #     raise XlsxExeption("Number of different columns")

        self.sheet.append(header)
        for line in data:
            self.sheet.append(line)
        self.book.save(self.workbook)
